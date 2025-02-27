from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import datetime
from sqlalchemy import func
from .. import db
from ..models import Workplace, Employee, Schedule, WorkplaceCost, WorkplaceRevenue, EmployeeCost, EmployeeRevenue
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import os

reports_bp = Blueprint('reports', __name__)

def format_currency(value):
    return f"{value:,.2f} zł"

def setup_header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

@reports_bp.route('/stats', methods=['POST'])
@jwt_required()
def get_statistics():
    try:
        data = request.get_json()
        start_date = datetime.fromisoformat(data.get('start_date').replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(data.get('end_date').replace('Z', '+00:00'))
        report_type = data.get('type', 'all')
        
        manager_id = get_jwt_identity()
        
        employees = Employee.query.filter_by(manager_id=manager_id).all()
        workplaces = Workplace.query.filter_by(manager_id=manager_id).all()

        employee_stats = []
        if report_type in ['employee', 'all']:
            for employee in employees:
                costs = EmployeeCost.query.filter(
                    EmployeeCost.employee_id == employee.id,
                    EmployeeCost.date.between(start_date, end_date)
                ).all()
                total_costs = sum(cost.amount for cost in costs)

                revenues = EmployeeRevenue.query.filter(
                    EmployeeRevenue.employee_id == employee.id,
                    EmployeeRevenue.date.between(start_date, end_date)
                ).all()
                total_revenues = sum(revenue.amount for revenue in revenues)

                schedules = Schedule.query.filter(
                    Schedule.employee_id == employee.id,
                    Schedule.date.between(start_date, end_date)
                ).all()
                total_hours = sum(schedule.hours for schedule in schedules)

                employee_stats.append({
                    'name': f'{employee.first_name} {employee.last_name}',
                    'total_costs': total_costs,
                    'total_revenues': total_revenues,
                    'total_profit': total_revenues - total_costs,
                    'total_hours': total_hours
                })
        
        workplace_stats = []
        if report_type in ['workplace', 'all']:
            for workplace in workplaces:
                costs = WorkplaceCost.query.filter(
                    WorkplaceCost.workplace_id == workplace.id,
                    WorkplaceCost.date.between(start_date, end_date)
                ).all()
                total_costs = sum(cost.amount for cost in costs)

                revenues = WorkplaceRevenue.query.filter(
                    WorkplaceRevenue.workplace_id == workplace.id,
                    WorkplaceRevenue.date.between(start_date, end_date)
                ).all()
                total_revenues = sum(revenue.amount for revenue in revenues)

                workplace_stats.append({
                    'name': f'{workplace.name}',
                    'total_costs': total_costs,
                    'total_revenues': total_revenues,
                    'total_profit': total_revenues - total_costs
                })
        
        return jsonify({
            'employees': employee_stats,
            'workplaces': workplace_stats
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/excel', methods=['POST'])
@jwt_required()
def generate_excel_report():
    user_id = get_jwt_identity()
    data = request.get_json()

    try:
        start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
        end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
        report_type = data.get('type', 'all')
    except (KeyError, ValueError):
        return jsonify({'error': 'Nieprawidłowy format dat'}), 400

    # Tworzenie pliku Excel
    wb = Workbook()
    
    try:
        if report_type in ['workplace', 'all']:
            ws_workplaces = wb.active
            ws_workplaces.title = "Miejsca pracy"
            
            headers_workplaces = [
                "Miejsce pracy",
                "Liczba pracowników",
                "Koszty",
                "Przychody",
                "Zysk"
            ]
            
            for col, header in enumerate(headers_workplaces, 1):
                cell = ws_workplaces.cell(row=1, column=col, value=header)
                setup_header_style(cell)
                ws_workplaces.column_dimensions[get_column_letter(col)].width = 15

            workplaces = Workplace.query.filter_by(owner_id=user_id).all()
            row = 2
            
            for workplace in workplaces:
                # Znajdź pracowników, którzy mają koszty lub przychody w tym miejscu pracy
                employees_with_activity = db.session.query(Employee)\
                    .join(WorkplaceRevenue, WorkplaceRevenue.employee_id == Employee.id, isouter=True)\
                    .filter(
                        WorkplaceRevenue.workplace_id == workplace.id,
                        WorkplaceRevenue.date.between(start_date, end_date)
                    )\
                    .distinct()\
                    .all()

                employee_count = len(employees_with_activity)

                # Oblicz koszty miejsca pracy
                costs = db.session.query(func.sum(WorkplaceCost.amount))\
                    .filter(
                        WorkplaceCost.workplace_id == workplace.id,
                        WorkplaceCost.date.between(start_date, end_date)
                    ).scalar() or 0

                # Oblicz przychody miejsca pracy
                revenues = db.session.query(func.sum(WorkplaceRevenue.amount))\
                    .filter(
                        WorkplaceRevenue.workplace_id == workplace.id,
                        WorkplaceRevenue.date.between(start_date, end_date)
                    ).scalar() or 0

                ws_workplaces.cell(row=row, column=1, value=workplace.name)
                ws_workplaces.cell(row=row, column=2, value=employee_count)
                ws_workplaces.cell(row=row, column=3, value=format_currency(float(costs)))
                ws_workplaces.cell(row=row, column=4, value=format_currency(float(revenues)))
                ws_workplaces.cell(row=row, column=5, value=format_currency(float(revenues - costs)))
                row += 1

        if report_type in ['employee', 'all']:
            # Generowanie arkusza dla pracowników
            ws_employees = wb.create_sheet("Pracownicy") if report_type == 'all' else wb.active
            if report_type == 'employee':
                ws_employees.title = "Pracownicy"
            
            headers_employees = [
                "Pracownik",
                "Miejsce pracy",
                "Koszty",
                "Przychody z miejsc pracy",
                "Przychody bezpośrednie",
                "Łączne przychody",
                "Zysk"
            ]
            
            for col, header in enumerate(headers_employees, 1):
                cell = ws_employees.cell(row=1, column=col, value=header)
                setup_header_style(cell)
                ws_employees.column_dimensions[get_column_letter(col)].width = 20

            employees = Employee.query\
                .filter_by(user_id=user_id)\
                .all()

            row = 2
            for employee in employees:
                # Oblicz koszty pracownika
                costs = db.session.query(func.sum(EmployeeCost.amount))\
                    .filter(
                        EmployeeCost.employee_id == employee.id,
                        EmployeeCost.date.between(start_date, end_date)
                    ).scalar() or 0

                # Oblicz przychody pracownika z miejsc pracy
                workplace_revenues = db.session.query(func.sum(WorkplaceRevenue.amount))\
                    .filter(
                        WorkplaceRevenue.employee_id == employee.id,
                        WorkplaceRevenue.date.between(start_date, end_date)
                    ).scalar() or 0

                # Oblicz bezpośrednie przychody pracownika
                direct_revenues = db.session.query(func.sum(EmployeeRevenue.amount))\
                    .filter(
                        EmployeeRevenue.employee_id == employee.id,
                        EmployeeRevenue.date.between(start_date, end_date)
                    ).scalar() or 0

                # Pobierz miejsca pracy pracownika
                workplaces = db.session.query(Workplace)\
                    .join(WorkplaceRevenue, WorkplaceRevenue.workplace_id == Workplace.id)\
                    .filter(
                        WorkplaceRevenue.employee_id == employee.id,
                        WorkplaceRevenue.date.between(start_date, end_date)
                    )\
                    .distinct()\
                    .all()

                workplace_names = ', '.join(wp.name for wp in workplaces) if workplaces else '-'
                total_revenues = workplace_revenues + direct_revenues

                ws_employees.cell(row=row, column=1, value=f"{employee.first_name} {employee.last_name}")
                ws_employees.cell(row=row, column=2, value=workplace_names)
                ws_employees.cell(row=row, column=3, value=format_currency(float(costs)))
                ws_employees.cell(row=row, column=4, value=format_currency(float(workplace_revenues)))
                ws_employees.cell(row=row, column=5, value=format_currency(float(direct_revenues)))
                ws_employees.cell(row=row, column=6, value=format_currency(float(total_revenues)))
                ws_employees.cell(row=row, column=7, value=format_currency(float(total_revenues - costs)))
                row += 1

        # Zapisz do pliku tymczasowego
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()

        # Wyślij plik
        return send_file(
            temp_file.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'raport_{start_date}_{end_date}.xlsx'
        )

    except Exception as e:
        # W przypadku błędu, upewnij się że plik tymczasowy zostanie usunięty
        if 'temp_file' in locals():
            os.unlink(temp_file.name)
        return jsonify({'error': 'Wystąpił błąd podczas generowania raportu Excel'}), 500

    finally:
        # Zawsze próbuj wyczyścić plik tymczasowy
        if 'temp_file' in locals():
            try:
                os.unlink(temp_file.name)
            except:
                pass 
